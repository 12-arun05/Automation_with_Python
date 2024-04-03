import openpyxl as xl
from openpyxl.chart import BarChart, LineChart, Reference

def process_worksheet(input_filename, output_filename, discount_percentage, graph_type, x_column, y_column):
    wb = xl.load_workbook(input_filename)
    sheet1 = wb['Sheet1']

    # Calculate the discounted prices
    discount_factor = 1 - (discount_percentage / 100)
    for row in range(2, sheet1.max_row + 1):
        cell = sheet1.cell(row, 4)
        discounted_price = cell.value * discount_factor
        discounted_price_cell = sheet1.cell(row, 6)
        discounted_price_cell.value = discounted_price

    # Set the title "Discounted price" for the 6th column (assuming first row contains headers)
    sheet1.cell(row=1, column=6).value = "Discounted price"

    # Get the data for the chart
    x_values = Reference(sheet1, min_row=2, max_row=sheet1.max_row, min_col=x_column, max_col=x_column)
    y_values = Reference(sheet1, min_row=2, max_row=sheet1.max_row, min_col=y_column, max_col=y_column)

    # Create the chart based on user's choice
    if graph_type == 'bar':
        chart = BarChart()
        chart.add_data(y_values)
        chart.set_categories(x_values)
    elif graph_type == 'line':
        chart = LineChart()
        chart.add_data(y_values)
        chart.set_categories(x_values)

    # Set x-axis and y-axis titles based on the selected columns
    x_axis_title = sheet1.cell(row=1, column=x_column).value
    y_axis_title = sheet1.cell(row=1, column=y_column).value
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title

    # Add the chart to the worksheet
    chart_cell = sheet1.cell(row=sheet1.max_row + 2, column=1)  # Adjust cell location as needed
    sheet1.add_chart(chart, chart_cell.coordinate)

    # Save the modified workbook
    wb.save(output_filename)

# Ask the user for inputs
input_file_name = input("Enter the name of the initial Excel file (with extension .xlsx): ")
output_file_name = input("Enter the name of the new Excel file to save (with extension .xlsx): ")
discount_percentage = float(input("Enter the discount percentage (e.g., 10 for 10%): "))
graph_type = input("Choose the type of graph (bar/line): ").lower()

# Validate user input for graph type
while graph_type not in ['bar', 'line']:
    print("Invalid input! Please choose either 'bar' or 'line'.")
    graph_type = input("Choose the type of graph (bar/line): ").lower()

x_column = int(input("Enter the column number for the x-axis data: "))
y_column = int(input("Enter the column number for the y-axis data: "))

# Call the function to process the worksheet and create the chart
process_worksheet(input_file_name, output_file_name, discount_percentage, graph_type, x_column, y_column)
